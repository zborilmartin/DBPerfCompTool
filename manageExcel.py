#!/usr/bin/env python
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color,PatternFill
from openpyxl.styles import colors
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl import load_workbook    
import os
from shutil import copyfile

# Setting colors and fonts
blueFill = PatternFill(start_color="1EB4F5",end_color="1EB4F5",fill_type='solid')
orangeFill = PatternFill(start_color="FFB011",end_color="FFB011",fill_type='solid')
yellowFill = PatternFill(start_color="FFF811",end_color="FFF811",fill_type='solid')
redFill = PatternFill(start_color="FF0000",end_color="FF0000",fill_type='solid')
greenFill = PatternFill(start_color="07EE1E",end_color="07EE1E",fill_type='solid')
darkredFill = PatternFill(start_color="990000",end_color="990000",fill_type='solid')
pinkFill = PatternFill(start_color="FF66B2",end_color="FF66B2",fill_type='solid')
lightblueFill = PatternFill(start_color="99FFFF",end_color="99FFFF",fill_type='solid')
brownFill = PatternFill(start_color="994C00",end_color="994C00",fill_type='solid')
lightgreenFill = PatternFill(start_color="CCFFCC",end_color="CCFFCC",fill_type='solid')
greyFill = PatternFill(start_color="C0C0C0",end_color="C0C0C0",fill_type='solid')
whiteFill = PatternFill(start_color="FFFFFF",end_color="FFFFFF",fill_type='solid')
noneFill = PatternFill(patternType=None,start_color=None,end_color=None,fill_type=None)
bold = Font(bold=True)
bold22 = Font(bold=True,size=22)
bold36 = Font(bold=True,size=36)

# Method for setting red or green fill -> is greater or smaller than particular cell
def formatRedGreenFill(ws,dbd,row_comparing,column_comparing,row_to_compare,column_to_compare,testname,tpch=0):
	if tpch == 0:
		ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_comparing, column=column_comparing).coordinate), CellIsRule(operator='equal', formula=["'DBD_{0}'!${1}".format(testname,dbd.cell(row=row_to_compare, column=column_to_compare).coordinate)], stopIfTrue=True, fill=orangeFill))
                ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_comparing, column=column_comparing).coordinate), CellIsRule(operator='greaterThan', formula=["'DBD_{0}'!${1}".format(testname,dbd.cell(row=row_to_compare, column=column_to_compare).coordinate)], stopIfTrue=True, fill=redFill))
                ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_comparing, column=column_comparing).coordinate), CellIsRule(operator='lessThan', formula=["'DBD_{0}'!${1}".format(testname,dbd.cell(row=row_to_compare, column=column_to_compare).coordinate)], stopIfTrue=True, fill=greenFill))
	else:
		ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_comparing, column=column_comparing).coordinate), CellIsRule(operator='equal', formula=["'DBD_{0}-ALL'!${1}".format(testname,dbd.cell(row=row_to_compare, column=column_to_compare).coordinate)], stopIfTrue=True, fill=orangeFill))
		ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_comparing, column=column_comparing).coordinate), CellIsRule(operator='greaterThan', formula=["'DBD_{0}-ALL'!${1}".format(testname,dbd.cell(row=row_to_compare, column=column_to_compare).coordinate)], stopIfTrue=True, fill=redFill))
		ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_comparing, column=column_comparing).coordinate), CellIsRule(operator='lessThan', formula=["'DBD_{0}-ALL'!${1}".format(testname,dbd.cell(row=row_to_compare, column=column_to_compare).coordinate)], stopIfTrue=True, fill=greenFill))  
	ws['{0}'.format(ws.cell(row=row_comparing, column=column_comparing).coordinate)].number_format = '### ### ### ### ###'

# Method for setting deferences on two metrics        
# Difference is situated one cell under comparing metrics
def formatDifference(ws,dbd,row_comparing,column_comparing,row_to_compare,column_to_compare,testname,tpch=0):
	if tpch == 0:
	        ws['{0}'.format(ws.cell(row=row_comparing+1, column=column_comparing).coordinate)] = "={0}/('DBD_{1}'!{2})".format(ws.cell(row=row_comparing, column=column_comparing).coordinate,testname,dbd.cell(row=row_to_compare, column=column_to_compare).coordinate)
		ws['{0}'.format(ws.cell(row=row_comparing+1, column=column_comparing).coordinate)].number_format = '0.###0'
	else:
		ws['{0}'.format(ws.cell(row=row_comparing+1, column=column_comparing).coordinate)] = "={0}/('DBD_{1}-ALL'!{2})".format(ws.cell(row=row_comparing, column=column_comparing).coordinate,testname,dbd.cell(row=row_to_compare, column=column_to_compare).coordinate)
		ws['{0}'.format(ws.cell(row=row_comparing+1, column=column_comparing).coordinate)].number_format = '0.###0'
	
def formatResult(ws,row_number, column_result):
	ws['{0}'.format(ws.cell(row=row_number, column=column_result).coordinate)] = "=({0}+{1}+{2})/3".format(ws.cell(row=row_number, column=column_result-5).coordinate,ws.cell(row=row_number, column=column_result-3).coordinate,ws.cell(row=row_number, column=column_result-2).coordinate)
	
	ws['{0}'.format(ws.cell(row=row_number, column=column_result).coordinate)].number_format = '0.###0'
	ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_number, column=column_result).coordinate), CellIsRule(operator='equal', formula=['1'], stopIfTrue=True, fill=orangeFill))
	ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_number, column=column_result).coordinate), CellIsRule(operator='greaterThan', formula=['1'], stopIfTrue=True, fill=redFill))
	ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_number, column=column_result).coordinate), CellIsRule(operator='lessThan', formula=['1'], stopIfTrue=True, fill=greenFill))


# Method for setting function AVERAGE for each query and COUNT of queries
def createAVGandCOUNT(ws1,column_start):
        for i in range (0,4):    
                ws1['{0}'.format(ws1.cell(row=9, column=column_start+i).coordinate)] = "=AVERAGE({0}:{1})".format(ws1.cell(row=501, column=column_start+4+i).coordinate,ws1.cell(row=10000, column=column_start+4+i).coordinate)
        ws1['{0}'.format(ws1.cell(row=9, column=column_start+4).coordinate)] = "=COUNT({0}:{1})".format(ws1.cell(row=501, column=column_start+4).coordinate,ws1.cell(row=10000, column=column_start+4).coordinate)
      
        
# Method for creating average table menu and table where the data are stored + formatting
def createAVGTable(ws1,column_start,testname):
	# Average - line
	ws1.cell(row=8, column=column_start, value="response_ms")
	ws1.cell(row=8, column=column_start+1, value="memory_allocated_kb")
	ws1.cell(row=8, column=column_start+2, value="memory_used_kb")
	ws1.cell(row=8, column=column_start+3, value="CPU_time")
	ws1.cell(row=8, column=column_start+4, value="queries_COUNT")
	ws1.cell(row=8, column=column_start+5, value="RESULT")
	ws1.cell(row=8, column=column_start+5).font=bold22 
    	# Formatting - blue color and bold 
    	for cellColumn in range(column_start,column_start+6):
        	ws1.cell(row=8,column=cellColumn).fill=blueFill
         	ws1.cell(row=8,column=cellColumn).font=bold
            
        # Formatting average table - data - orange fill
        for cellColumn in range(column_start,column_start+5):
            if ws1.title == "DBD_{0}".format(testname):
        	       ws1.cell(row=9,column=cellColumn).fill=orangeFill
   	    ws1.cell(row=9,column=cellColumn).number_format = '### ### ### ### ###'
          
	# Items - line
    #	ws1.cell(row=499, column=column_start, value=int(0))
	ws1["{0}".format(ws1.cell(row=499, column=column_start).coordinate)]=0
	ws1["{0}".format(ws1.cell(row=499, column=column_start+1).coordinate)]="=COUNT({0}:{1})".format(ws1.cell(row=501, column=column_start+4).coordinate,ws1.cell(row=10000, column=column_start+4).coordinate)  
	ws1.cell(row=500, column=column_start, value="start_timestamp")
	ws1.cell(row=500, column=column_start+1, value="end_timestamp")
	ws1.cell(row=500, column=column_start+2, value="transaction_id")
	ws1.cell(row=500, column=column_start+3, value="statement_id")
	ws1.cell(row=500, column=column_start+4, value="response_ms")
	ws1.cell(row=500, column=column_start+5, value="memory_allocated_kb")
	ws1.cell(row=500, column=column_start+6, value="memory_used_kb")
	ws1.cell(row=500, column=column_start+7, value="CPU_time")
    	ws1.cell(row=500, column=column_start+8, value="Label/Query")
        ws1.cell(row=500, column=column_start+9, value="Table schema")

    	# Formatting of table where data are stored
    	for cellColumn in range(column_start,column_start+9):
        	ws1.cell(row=500,column=cellColumn).fill=yellowFill
            	ws1.cell(row=500,column=cellColumn).font=bold
        
# Method for creating Overview - Basic page - only Head and DBD
def createOverview(ws,queries,testname):
    	# Header
        ws.cell(row=1, column=1, value="Comparison - Projections of Vertica Database Design x Own projections")
    	ws.cell(row=1, column=1).font = bold36
    	
        # Testname
	ws.cell(row=2, column=1, value="Testname:")
	ws.cell(row=2, column=2, value=testname)
        ws.cell(row=2, column=2).font = bold22
        
        # Queries
	ws.cell(row=3, column=1, value="Queries:")
	i = 2
	for query in queries:
		ws.cell(row=3, column=i, value=query)
        	ws.cell(row=1, column=i).font = bold22
		i += 1

        # Number of schemas
    	ws.cell(row=4, column=1, value="Number of schemas:")
	ws.cell(row=4, column=2, value=1)  
        
        # Schema - DBD
	ws.cell(row=8, column=1, value="Schema:")
    	ws.cell(row=8, column=2, value="DBD")
    	ws.cell(row=8, column=2).font=bold36
    
        # Description - DBD
	ws.cell(row=8, column=14, value="Description:")
    
        # Copying AVG table from DBD schema
        for part in range(1,len(queries)+2):
        # There is 5 columns: time,time,used memory, cpu, count of queries
                ##for i in range (1,6):
                row_start = 8
                        # Description
                ws.cell(row=row_start, column=14, value="Description:")
                for i in range (1,6):
                        column_one = (part-1)*10 + i
                        # Copying data
                        for j in range (0,2):
                                ws['{0}'.format(ws.cell(row=row_start+j+1, column=column_one).coordinate)] = "='DBD_{0}'!{1}".format(testname,ws.cell(row=8+j, column=column_one).coordinate)
                        ws.cell(row=row_start+1,column=column_one).fill=blueFill
                        ws.cell(row=row_start+1,column=column_one).font=bold
			ws.cell(row=row_start+2,column=column_one).fill=orangeFill			
			ws.cell(row=row_start+2,column=column_one).number_format = '### ### ### ### ###'

		if part != 1:
			ws.cell(row=row_start+1,column=(part-1)*10).font=bold22
			ws.cell(row=row_start+1,column=(part-1)*10).value = '={0}'.format(ws.cell(row=3,column=part).coordinate)

	for i in range (1,10):
                row_start = 15
                ws.cell(row=row_start, column=1, value="Projection - Bytes:")
                ws.cell(row=row_start,column=1).fill=lightblueFill
                ws.cell(row=row_start,column=1).font=bold
                ws.cell(row=row_start+1,column=i).fill=lightblueFill
                ws.cell(row=row_start+1,column=i).font=bold
                for j in range (0,2):
                        ws['{0}'.format(ws.cell(row=row_start+j+1, column=i).coordinate)] = "='DBD_{0}'!{1}".format(testname,ws.cell(row=13+j, column=i).coordinate)
		ws.cell(row=row_start+2,column=i).number_format = '### ### ### ### ###'

        ws.cell(row=15, column=11, value='All TPC-H queries')
        ws.cell(row=15, column=11).font=bold22

        for i in range (1,6):
                column_one = 10 + i
                # Copying data
                row_start = 15
                for j in range (0,2):
                        ws['{0}'.format(ws.cell(row=row_start+j+1, column=column_one).coordinate)] = "='DBD_{0}-ALL'!{1}".format(testname,ws.cell(row=8+j, column=i).coordinate)
                ws.cell(row=row_start+1,column=column_one).fill=blueFill
                ws.cell(row=row_start+1,column=column_one).font=bold
		ws.cell(row=row_start+2,column=column_one).fill=orangeFill

	column_widths = []
	for row in ws.rows:
	    for i, cell in enumerate(row):
		if len(column_widths) > i:
		    if len(cell) > column_widths[i]:
			column_widths[i] = len(cell)

	for i, column_width in enumerate(column_widths):
	    worksheet.column_dimensions[get_column_letter(i+1)].width = column_width



	# Width of cells    
#	dims = {}
#	for row in ws.rows:
#		for cell in row:
#			if cell.value:
#				dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
#	for col, value in dims.items():
#		ws.column_dimensions[col].width = value    
        
# Method for adding new schema to Overview
def addToOverview(wb,ws,new,queries,schema,testname):  
        # DBD schema - for comparing
    	dbd = wb['DBD_{0}'.format(testname)]
	name = schema + '-ALL'
	newAll = wb[name]
	dbdAll = wb['DBD_{0}-ALL'.format(testname)]
    	# For appropriate number of queries (>1 query >>> 1 more table)    
	for part in range(1,len(queries)+2):
        # There is 5 columns: time,time,used memory, cpu, count of queries
		##for i in range (1,6):
            	row_start = 21 + (int(ws.cell(row=4,column=2).value)-1)*10
			##column_one = (part-1)*10 + i
                        # Schema and Description
	        ws.cell(row=row_start, column=1, value="Schema:")
    	    	ws.cell(row=row_start, column=2, value=schema)
            	ws.cell(row=row_start, column=2).font=bold36
    
    		        # Description
	       	ws.cell(row=row_start, column=14, value="Description:")
            	for i in range (1,6):
			column_one = (part-1)*10 + i
			# Copying data
			for j in range (0,3):
				if i == 5 and j == 2:
                                	continue
				ws['{0}'.format(ws.cell(row=row_start+j+1, column=column_one).coordinate)] = "={0}!{1}".format(schema,ws.cell(row=8+j, column=column_one).coordinate)
			ws.cell(row=row_start+1,column=column_one).fill=blueFill
			ws.cell(row=row_start+1,column=column_one).font=bold

					# Formatting - green/red 
			if i < 5: 
				formatRedGreenFill(ws,ws,row_start+2,column_one,9,column_one,testname)
				ws.cell(row=row_start+3,column=column_one).number_format = '0.###0'

		column_start = (part-1)*10 + 1
		formatResult(ws,row_start+3,column_start + 5)
	
		ws.cell(row=row_start+1,column=(part-1)*10 + 6).fill=blueFill
                ws.cell(row=row_start+1,column=(part-1)*10 + 6).font=bold
		ws.cell(row=row_start+1,column=(part-1)*10 + 6, value = 'Result')
                if part != 1:
                        ws.cell(row=row_start+1,column=(part-1)*10).font=bold22
                        ws.cell(row=row_start+1,column=(part-1)*10).value = '={0}'.format(ws.cell(row=3,column=part).coordinate)

	for i in range (1,6):
		column_one = 10 + i
		# Copying data
		row_start = 26 + (int(ws.cell(row=4,column=2).value)-1)*10
		if i == 1:
			ws.cell(row=row_start, column=column_one, value='All TPC-H queries')
        		ws.cell(row=row_start, column=column_one).font=bold22
		for j in range (0,3):
			if i == 5 and j == 3:
				break
			ws['{0}'.format(ws.cell(row=row_start+j+1, column=column_one).coordinate)] = "='{0}-ALL'!{1}".format(schema,newAll.cell(row=8+j, column=i).coordinate)
		ws.cell(row=row_start+1,column=column_one).fill=blueFill
		ws.cell(row=row_start+1,column=column_one).font=bold
		ws.cell(row=row_start+3,column=column_one).number_format = '0.###0'

		if i < 5 :		# Formatting - green/red 
			formatRedGreenFill(ws,dbdAll,row_start+2,column_one,9,i,testname,1)
        formatResult(ws,26 + (int(ws.cell(row=4,column=2).value)-1)*10+3,16)


	for i in range (1,10):
        	row_start = 26 + (int(ws.cell(row=4,column=2).value)-1)*10
		ws.cell(row=row_start, column=1, value="Projection - Bytes:")
		ws.cell(row=row_start,column=1).fill=lightblueFill
                ws.cell(row=row_start,column=1).font=bold
                column_one = (part-1)*10 + i
		ws.cell(row=row_start+1,column=i).fill=lightblueFill
                ws.cell(row=row_start+1,column=i).font=bold
		for j in range (0,3):
	                ws['{0}'.format(ws.cell(row=row_start+j+1, column=i).coordinate)] = "={0}!{1}".format(schema,ws.cell(row=13+j, column=i).coordinate)
			formatRedGreenFill(ws,ws,row_start+2,i,14,i,testname)
		ws.cell(row=row_start+2,column=i).number_format = '0.###0'
	
	row_start = 26 + (int(ws.cell(row=4,column=2).value)-1)*10
	ws.cell(row=row_start+1,column=16).fill=blueFill
        ws.cell(row=row_start+1,column=16).font=bold
        ws.cell(row=row_start+1,column=16, value = 'Result')	

		# Width of cells
	dims = {}
	for row in ws.rows:
		for cell in row:
			if cell.value:
				dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
	for col, value in dims.items():
		ws.column_dimensions[col].width = value

def createProfile(ws1,column_start):
        # Query profile
	ws1.cell(row=17, column=column_start, value="Explain Verbose")
        ws1.cell(row=17,column=column_start).fill=blueFill
        ws1.cell(row=17,column=column_start+1).fill=blueFill
        ws1.cell(row=17,column=column_start).font=bold22        

	# Query profile - line
	#ws1.cell(row=18, column=column_start, value="running_time")
	#ws1.cell(row=18, column=column_start+1, value="memory_allocated_bytes")
	#ws1.cell(row=18, column=column_start+2, value="read_from_disk_bytes")
	#ws1.cell(row=18, column=column_start+3, value="path_line")    
 	#for cellColumn in range(column_start,column_start+5):
       # 	ws1.cell(row=18,column=cellColumn).fill=blueFill
       #  	ws1.cell(row=18,column=cellColumn).font=bold           
  
        # Size of projections
	ws1.cell(row=12,column=1, value="Projection - Bytes:")
        ws1.cell(row=12,column=1).fill=lightblueFill    
        ws1.cell(row=12,column=1).font=bold
        i = 0
        for name in ["Customer:","Lineitem:","Nation:","Orders:","Part:","Partsupp:","Region:","Supplier:","SUM:"]:
                ws1.cell(row=13,column=1+i, value=name)
                ws1.cell(row=13,column=1+i).fill=lightblueFill
		ws1.cell(row=14,column=1+i).number_format = '### ### ### ### ###'
                #ws1.cell(row=14,column=1+i).fill=greyFill
                i += 1

        ws1.cell(row=13,column=column_start+8).font=bold
        ws1['{0}'.format(ws1.cell(row=14, column=9).coordinate)] = "=SUM({0}:{1})".format(ws1.cell(row=14, column=1).coordinate,ws1.cell(row=14, column=8).coordinate)

def loadDataToExcelToParticularTable(row,ws,start_column):
	#ws.cell(row=499,column=start_column).value = 0
	ws.cell(row=499,column=start_column).value += int(1)
	for i in range (1,10):
    		# integer data type, 1.column is schema - not storing in Excel file
		if i in [5,6,7,8]:
			ws['{0}'.format(ws.cell(row=500+ws.cell(row=499,column=start_column).value, column=start_column+i-1).coordinate)] = int(row[i])
    		# other data type
		else:   
			ws['{0}'.format(ws.cell(row=500+ws.cell(row=499,column=start_column).value, column=start_column+i-1).coordinate)] = row[i]    
	ws['{0}'.format(ws.cell(row=500+ws.cell(row=499,column=start_column).value, column=start_column+10-1).coordinate)] = row[0]


def loadDataToExcel(rows,query,schema,testname,queries,tpch=0):
        # Opening excel - must be created and sheet with specific schema must be created
        wb = load_workbook('CompareOutput/{0}.xlsx'.format(testname))
        ws = wb[schema]
	for row in rows:
        	start_column = 1
            	#8 columns to store
	        loadDataToExcelToParticularTable(row,ws,start_column)
		if tpch == 1:
			continue
        	## increasing number of queries
        	# If there is more than 1 query, data are stored also in particular query table
		if len(queries)>1:
	                index = queries.index(query)
			start_column_query= 10*(int(index)+1) + 1
			loadDataToExcelToParticularTable(row,ws,start_column_query)
        wb.save('CompareOutput/' + testname + '.xlsx')

# Method for creating new sheer according to new schema
# Particular step: duplicate Pattern sheet and fill Schema name
def duplicatePattern(schema,testname,queries,query):
    wb = load_workbook('CompareOutput/{0}.xlsx'.format(testname))
    # Creating snew sheet
    if schema not in wb.get_sheet_names():
        ws = wb["DBD_{0}".format(testname)]
        wsAll = wb["DBD_{0}-ALL".format(testname)]
        overview = wb["Overview"]
        pattern = wb["Pattern"]
        patternAll = wb["Pattern-ALL"]

        new = wb.copy_worksheet(pattern)
        new.cell(row=3, column=2, value=schema)
        new.title = schema

        newAll = wb.copy_worksheet(patternAll)
	name = schema + ' - All TPC-H queries'
        newAll.cell(row=3, column=2, value=name)
        newAll.title = schema + '-ALL'

        addToOverview(wb,overview,new,queries,schema,testname)
        overview.cell(row=4,column=2).value = int(overview.cell(row=4,column=2).value) + 1 

	for part in range(1,len(queries)+2):

		for i in range(1,5):
			formatRedGreenFill(new,ws,9,((part-1)*10 + i),9,((part-1)*10 + i),testname)

		for i in range (1,5):
			formatDifference(new,ws,9,((part-1)*10+i),9,((part-1)*10+i),testname)

		formatResult(new,10,((part-1)*10 + 6))

		if len(queries) == 1:
			break
		#row_number = 10
		#column_result = (part-1)*10 + 6
		#ws['{0}'.format(ws.cell(row=row_number, column=column_result).coordinate)].number_format = '0.###0'
        	#ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_number, column=column_result).coordinate), CellIsRule(operator='equal', formula=['1'], stopIfTrue=True, fill=orangeFill))
        	#ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_number, column=column_result).coordinate), CellIsRule(operator='greaterThan', formula=['1'], stopIfTrue=True, fill=redFill))
        	#ws.conditional_formatting.add('{0}'.format(ws.cell(row=row_number, column=column_result).coordinate), CellIsRule(operator='lessThan', formula=['1'], stopIfTrue=True, fill=greenFill))

	#	if len(queries) > 1 and part == 1:
	#		continue
		

		# USED BYTES PROJECTIONS
	for j in range(1,10):
		formatRedGreenFill(new,ws,14,j,14,j,testname)
		formatDifference(new,ws,14,j,14,j,testname)	
	# Setting format for Query Profile Plan part of Excel file    
        new = formatQueryProfilePlan(new)

        for i in range(1,4):
                formatRedGreenFill(newAll,wsAll,9,i,9,i,testname,1)
		formatDifference(newAll,wsAll,9,i,9,i,testname,1)

                # USED BYTES PROJECTIONS
        for j in range(1,10):
                formatRedGreenFill(newAll,wsAll,14,j,14,j,testname,1)
                formatDifference(newAll,wsAll,14,j,14,j,testname,1) 

	wb.save('CompareOutput/' + testname + '.xlsx')        

# Method for loading Query Profile Path into the excel file        
def loadExplain(schema,testname,rows,queries,query,monitor=0):
	wb = load_workbook('CompareOutput/{0}.xlsx'.format(testname))
    	ws = wb[schema]
	#name = schema + '-ALL'	
	name = schema
	wsAll = wb[name]
	start_column = 1
	if len(queries) > 1:
		index = queries.index(query)                
		start_column= 10*(int(index)+1) + 1
	tmp_row = 0
	if monitor == 0:
		for row in rows:
			tmp_column = 0
			for item in row:
				ws.cell(row=19+tmp_row,column=start_column+tmp_column,value=item)
				tmp_column += 1
			tmp_row += 1
	if schema[-4:] == '-ALL':
		schema = schema[:-4]
	fileSize = './ExplainProfile/{0}/Projection_size_{1}_{2}.txt'.format(testname,testname,schema)
	sizeFile = open(fileSize, 'r+')	
	for i in range(0,9):
		f = sizeFile.readline().split('\n')
		column_index = 1 + i
		ws.cell(row=14,column=column_index,value=int(f[0])) 
		wsAll.cell(row=14,column=column_index,value=int(f[0])) 
	sizeFile.close()
	wb.save('CompareOutput/{0}.xlsx'.format(testname))

         
def formatQueryProfilePlan(ws1):  
	dxf = DifferentialStyle(fill=yellowFill)
	rule = Rule(type="containsText", operator="containsText", text="> JOIN", dxf=dxf)
	rule.formula = ['NOT(ISERROR(SEARCH("> JOIN",A19)))']
	ws1.conditional_formatting.add('A19:ZZ499', rule)

	dxf = DifferentialStyle(fill=lightblueFill)
	rule = Rule(type="containsText", operator="containsText", text="Filter", dxf=dxf)
	rule.formula = ['NOT(ISERROR(SEARCH("Filter",A19)))']
	ws1.conditional_formatting.add('A19:ZZ499', rule)       
	
	dxf = DifferentialStyle(fill=orangeFill)
	rule = Rule(type="containsText", operator="containsText", text="Join Cond", dxf=dxf)
	rule.formula = ['NOT(ISERROR(SEARCH("Join Cond",A19)))']
	ws1.conditional_formatting.add('A19:ZZ499', rule)
	
	dxf = DifferentialStyle(fill=pinkFill)
	rule = Rule(type="containsText", operator="containsText", text="Projection:", dxf=dxf)
	rule.formula = ['NOT(ISERROR(SEARCH("Projection:",A19)))']
	ws1.conditional_formatting.add('A19:ZZ499', rule)
	
	dxf = DifferentialStyle(fill=lightgreenFill)
	rule = Rule(type="containsText", operator="containsText", text="SELECT", dxf=dxf)
	rule.formula = ['NOT(ISERROR(SEARCH("SELECT",A19)))']
	ws1.conditional_formatting.add('A19:ZZ499', rule)
	
	dxf = DifferentialStyle(fill=darkredFill)
	rule = Rule(type="containsText", operator="containsText", text="SORT [", dxf=dxf)
	rule.formula = ['NOT(ISERROR(SEARCH("SORT [",A19)))']
	ws1.conditional_formatting.add('A19:ZZ499', rule)
	
	dxf = DifferentialStyle(fill=redFill)
	rule = Rule(type="containsText", operator="containsText", text="> GROUPBY", dxf=dxf)
	rule.formula = ['NOT(ISERROR(SEARCH("> GROUPBY",A19)))']
	ws1.conditional_formatting.add('A19:ZZ499', rule)
	
	dxf = DifferentialStyle(fill=blueFill)
	rule = Rule(type="containsText", operator="containsText", text="Outer -> STORAGE", dxf=dxf)
	rule.formula = ['NOT(ISERROR(SEARCH("Outer -> STORAGE",A19)))']
	ws1.conditional_formatting.add('A19:ZZ499', rule)
	
	dxf = DifferentialStyle(fill=greenFill)
	rule = Rule(type="containsText", operator="containsText", text="Inner -> STORAGE", dxf=dxf)
	rule.formula = ['NOT(ISERROR(SEARCH("Inner -> STORAGE",A19)))']
	ws1.conditional_formatting.add('A19:ZZ499', rule)
	# end 	
	
	return ws1
  
# Method for creating Excel file if the file is not created already
def createExcelFile(testname,queries):
        if not os.path.exists('./CompareOutput'):
        	os.makedirs('./CompareOutput')

    	if not os.path.exists('./CompareOutput/' + testname + '.xlsx'):
		# Create workbook and sheets
		wb = Workbook()
		ws = wb.active
        	# There is 2 sheets - Overview (dashboard) and DBD
		ws1 = wb.create_sheet("DBD_{0}".format(testname))
		ws.title = "Overview"

		# Header
		    # Testname
		ws1.cell(row=1, column=1, value="Testname:")
		ws1.cell(row=1, column=2, value=testname)
		ws1.cell(row=1, column=2).font = bold22
		
		    # Queries
		ws1.cell(row=2, column=1, value="Queries:")
		i = 2
		for query in queries:
			ws1.cell(row=2, column=i, value=query)
			ws1.cell(row=2, column=i).font = bold22
			i += 1
		    # Schema and Description
		ws1.cell(row=3, column=1, value="Schema:")
		ws1.cell(row=3, column=2).font=bold36
	    
		    # Description
		ws1.cell(row=4, column=1, value="Description:")
	    
		    # Overview
		if len(queries)>1:
			ws1.cell(row=5, column=1, value="OVERVIEW")
			ws1.cell(row=5, column=1).fill=yellowFill
			ws1.cell(row=5, column=1).font=bold22

		# Table lines   
		tmp = 0
		for part in range(1,len(queries)+2):
			createAVGTable(ws1,(part-1)*10 + 1,testname)
			createAVGandCOUNT(ws1,(part-1)*10 + 1)
			if len(queries) == 1:
				createProfile(ws1,1)
				break			                   
			if part > 1:
				createProfile(ws1,(part-1)*10 + 1)                        
				ws1.cell(row=5,column=(part-1)*10 + 1,value="Query: ")
				ws1.cell(row=5,column=(part-1)*10 + 2,value=str(queries[part-2]))
				ws1.cell(row=5,column=(part-1)*10 + 2).font=bold22
			
        	# Setting format for Query Profile Plan part of Excel file    
		ws1 = formatQueryProfilePlan(ws1)				    
	
		# Width of cells
		dims = {}
		for row in ws1.rows:
			for cell in row:
				if cell.value:
					dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
		for col, value in dims.items():
			ws1.column_dimensions[col].width = value

				    
		# Create pattern for other schemas
		pattern = wb.copy_worksheet(ws1)
       		
		pattern.title = "Pattern"
                        
                # Create pattern for ALL TPCH queries
                patternAll = wb.copy_worksheet(ws1)
		patternAll.title = "Pattern-ALL"

		for row in patternAll.iter_rows('{0}:{1}'.format(patternAll.cell(row=5, column=11).coordinate,patternAll.cell(row=500, column=((len(queries)+1)*10)).coordinate)):
                        #if len(queries) == 1:
                        #        break
			for cell in row:
				cell.value = None			
				cell.fill=noneFill

		wsAll = wb.copy_worksheet(patternAll)
                wsAll.title = "DBD_{0}-ALL".format(testname)

        	# In DBD sheet set schema name to DBD
		ws1.cell(row=3, column=2, value="DBD_{0}".format(testname))
		wsAll.cell(row=3, column=2, value="DBD_{0}".format(testname))
	    
		# Formatting average table - green/red - in Pattern sheet
		for part in range(1,len(queries)+2):			
			for i in range (1,5):
                		formatDifference(pattern,ws1,9,((part-1)*10+i),9,((part-1)*10+i),testname)
			
			formatResult(pattern,10,((part-1)*10 + 6))		
	   
			if len(queries) == 1:
				break
				

		# USED BYTES PROJECTIONS
            	for j in range(1,10):
                    	formatDifference(pattern,ws1,14,j,14,j,testname)

		for i in range (1,5):
                	formatDifference(patternAll,wsAll,9,i,9,i,testname,1)
		formatResult(patternAll,10,6)

		for j in range(1,10):
                        formatDifference(patternAll,wsAll,14,j,14,j,testname,1)


        	# Method for creating Overview        
		createOverview(ws,queries,testname) 
 
		wb.save('CompareOutput/' + testname + '.xlsx')
	else:
		if not os.path.exists('./CompareOutput/Backups'):
                	os.makedirs('./CompareOutput/Backups')
		copyfile('./CompareOutput/' + testname + '.xlsx','./CompareOutput/Backups/' + testname + '_backup.xlsx')
